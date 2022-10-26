import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}

total_value_of_products = {}

products_inv_under_10 = {}

# Lists each company with respective product count
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value  # brings value in 4th cell in product_row
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_product_num = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_product_num + 1
    else:
        print("Adding a new supplier..")
        products_per_supplier[supplier_name] = 1

    # Total value of products per supplier
    if supplier_name in total_value_of_products:
        current_value = total_value_of_products.get(supplier_name)
        total_value_of_products[supplier_name] = current_value + inventory * price
    else:
        total_value_of_products[supplier_name] = inventory * price

    # Logic to find products in inventory under amount of 10
    if inventory < 10:
        products_inv_under_10[product_num] = inventory

    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")
print(f"Products per supplier: {products_per_supplier}")
print(f"Total value of the products per supplier: {total_value_of_products}")
print(f"Products need to be reordered: {products_inv_under_10}")
